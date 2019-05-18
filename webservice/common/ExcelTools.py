# -*- coding=utf-8 -*-
import openpyxl
from common import constant


class Case:

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.method = None
        self.data = None
        self.expected = None
        self.check_sql = None


class DoExcel:

    def __init__(self, files, sheet_name):
        self.files = files
        self.sheet_name = sheet_name
        self.lw = openpyxl.load_workbook(files)
        self.sheet = self.lw[sheet_name]

    def read_excel(self):
        cases = []
        for row in range(2, self.sheet.max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(row, 1).value
            case.title = self.sheet.cell(row, 2).value
            case.url = self.sheet.cell(row, 3).value
            case.method = self.sheet.cell(row, 4).value
            case.data = self.sheet.cell(row, 5).value
            case.expected = self.sheet.cell(row, 6).value
            case.check_sql = self.sheet.cell(row, 9).value
            cases.append(case)
        self.lw.close()
        return cases

    def write_excel(self, case_id, actual, result):
        sheet = self.lw[self.sheet_name]
        sheet.cell(case_id + 1, 7).value = actual
        sheet.cell(case_id + 1, 8).value = result
        self.lw.save(self.files)
        self.lw.close()


if __name__ == '__main__':
    do_excel = DoExcel(constant.excel_dir, 'sendmcode')
    do_excel.write_excel(1, 'testcase', 'pase')
